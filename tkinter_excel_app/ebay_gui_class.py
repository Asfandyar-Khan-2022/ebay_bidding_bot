"""eBay bidding bot GUI integrated with bidding."""

import tkinter as tk
from tkinter import ttk
import openpyxl
from bidder_class import EbayBidding
from apscheduler.schedulers.background import BackgroundScheduler
from datetime import datetime as dt
import datetime


class EbayGui():
    """GUI integrated with bidder to bid on items."""
    def __init__(self):
        """Select Tkinter style and start scheduler.

        Args:
            root (Tk)                       : Creates the GUI.
            style (Style)                   : Sets the GUI theme.
            combo_list (list)               : Range of seconds.
            scheduler (BackgroundScheduler) : Run scheduler in background.

        """
        self.root = tk.Tk()
        self.style = ttk.Style(self.root)
        self.root.tk.call("source", "forest-dark.tcl")
        self.style.theme_use("forest-dark")
        self.combo_list = [5, 4, 3, 2, 1]
        self.scheduler = BackgroundScheduler()
        self.scheduler.start()
        

    def load_path(self, path: str) -> str:
        """Loads the excel file.

        Args:
            path: path to the name of excel to open

        Returns:
            The contents of the excel

        """
        workbook = openpyxl.load_workbook(path)
        return workbook.active

    def set_frame(self) -> None:
        """Creates the GUI window"""
        self.frame = ttk.Frame(self.root)
        self.frame.pack()

    def bid_border(self) -> None:
        """Create bidding bot section within GUI"""
        self.widgets_frame = ttk.LabelFrame(self.frame, text="Insert Bid")
        self.widgets_frame.grid(row=0, column=0, sticky="nsew")

    def id_row(self) -> None:
        """GUI box to add item ID"""
        self.item_id = ttk.Entry(self.widgets_frame)
        self.item_id.insert(0, "Item ID")
        self.item_id.bind(
            "<FocusIn>",
            lambda e: self.item_id.delete("0", "end"))
        self.item_id.grid(row=0, column=0, padx=5, pady=10, sticky="ew")

    def amount_row(self) -> None:
        """GUI box to add item amount"""
        self.amount = ttk.Entry(self.widgets_frame)
        self.amount.insert(0, "£")
        self.amount.bind("<FocusIn>", lambda e: self.amount.delete("0", "end"))
        self.amount.grid(row=1, column=0, padx=5, pady=10, sticky="ew")

    def seconds_row(self) -> None:
        """GUI dropdown to add item ID"""
        self.seconds_combobox = ttk.Combobox(
            self.widgets_frame,
            values=self.combo_list)
        self.seconds_combobox.current(0)
        self.seconds_combobox.grid(
            row=2,
            column=0,
            padx=5,
            pady=10,
            sticky="ew")

    def load_chrome_path(self) -> list:
        """Get excel data.

        Returns:
            List of chrome paths

        """
        sheet = self.load_path("chrome_path.xlsx")
        return list(sheet.values)[1:]

    def default_or_user_path(self) -> list:
        """Set user inserted path or default values.

        Returns:
            User path if possible, default otherwise.

        """
        if None in self.load_chrome_path()[1]:
            return self.load_chrome_path()[0]
        else:
            return self.load_chrome_path()[1]

    def chrome_tkinter_setup(self, list_values: str, row: int) -> ttk.Entry:
        """Setup row to insert chrome path data.

        Args:
            list_values: Chrome path.
            row: Row to place the value into.

        Returns:
            Row setup settings

        """
        setup = ttk.Entry(self.frame)
        setup.insert(0, f"{list_values}")
        setup.bind("<FocusIn>", lambda e: setup.delete("0", "end"))
        setup.grid(
            row=row,
            column=0,
            columnspan=2,
            padx=5,
            pady=10,
            sticky="ew")
        return setup

    def set_chrome_profile(self) -> None:
        """Create row to insert chrome profile."""
        list_values = self.default_or_user_path()[0]
        self.user_profile = self.chrome_tkinter_setup(
            list_values=list_values,
            row=1)

    def set_chrome_user_path(self) -> None:
        """Create row to insert chrome user path."""
        list_values = self.default_or_user_path()[1]
        self.user_path = self.chrome_tkinter_setup(
            list_values=list_values,
            row=2)

    def set_chrome_exe(self) -> None:
        """Create row to insert chrome executable."""
        list_values = self.default_or_user_path()[2]
        self.set_exe = self.chrome_tkinter_setup(
            list_values=list_values,
            row=3)

    def save_chrome_path_and_exe(self) -> None:
        """Save the inserted chrome path, exe and profile to excel."""
        path = "chrome_path.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        sheet["A3"] = self.user_profile.get()
        sheet["B3"] = self.user_path.get()
        sheet["C3"] = self.set_exe.get()
        workbook.save(path)

    def save_path(self) -> None:
        """Submit button to save chrome path, exe and profile."""
        save_path = tk.Button(
            self.widgets_frame,
            text="Save chrome path",
            command=self.save_chrome_path_and_exe)

        save_path.grid(row=6, column=0, padx=5, pady=10, sticky="nsew")

    def submit_bid(self) -> None:
        """Submit button to place bid."""
        button = ttk.Button(
            self.widgets_frame,
            text="Bid",
            command=self.type_error)
        button.grid(row=4, column=0, padx=5, pady=10, sticky="nsew")

    def error_log(self) -> None:
        """Shows any errors made in block."""
        self.errors_frame = ttk.LabelFrame(
            self.frame,
            text="Errors Log",
            width=100,
            height=100)
        self.errors_frame.grid(row=4, column=0, columnspan=2, sticky="nsew")

    def error_text(self) -> None:
        """Label error log block."""
        self.error = tk.Label(self.errors_frame, text="")
        self.error.pack(pady=20)

    def invalid_id_error(self) -> None:
        """Attempts to get valid item ID."""
        try:
            int(self.item_id.get())
        except ValueError:
            self.error.config(text="Please enter a valid item ID. It should "
                              "look something like this: 101202921365")

    def invalid_amount_error(self) -> None:
        """Attempts to get valid item ID."""
        try:
            float(self.amount.get())
        except ValueError:
            self.error.config(text="Please enter a valid amount. "
                              "It should look somethig like this: 10 or 10.5")

    def invalid_seconds_error(self) -> None:
        """Attempts to get valid time in seconds."""
        try:
            int(self.seconds_combobox.get())
        except ValueError:
            self.error.config(text="Please enter a valid time. "
                              "It should look something like this: 3")

    def insert_valid_row(self) -> None:
        """Attempts to get item ID, amount and seconds."""
        try:
            int(self.item_id.get())
            float(self.amount.get())
            int(self.seconds_combobox.get())
            self.error.config(text=" ")
            self.insert_row()
        except ValueError:
            pass

    def type_error(self) -> None:
        """Ordered method call."""
        self.invalid_id_error()
        self.invalid_amount_error()
        self.invalid_seconds_error()
        self.insert_valid_row()

    def bid_history(self) -> None:
        """Create column to add bid queued items."""
        self.treeFrame = ttk.Frame(self.frame)
        self.treeFrame.grid(row=0, column=1)
        self.treeScroll = ttk.Scrollbar(self.treeFrame)
        self.treeScroll.pack(side="right", fill="y")

    def load_history(self) -> None:
        """Add item rows to bid history column."""
        cols = ("Item ID", "Amount £", "Seconds", "Bid end")
        self.treeview = ttk.Treeview(
            self.treeFrame, show="headings",
            yscrollcommand=self.treeScroll.set,
            columns=cols,
            height=13)

        self.treeview.column("Item ID", width=100)
        self.treeview.column("Amount £", width=50)
        self.treeview.column("Seconds", width=100, anchor="center")
        self.treeview.column("Bid end", width=200)
        self.treeview.pack()
        self.treeScroll.config(command=self.treeview.yview)
        self.load_data()

    def delete_row(self) -> None:
        """Delete item row from bid queue."""
        delete_button = tk.Button(
            self.widgets_frame,
            text="Delete row",
            command=lambda: self.delete())

        delete_button.grid(row=5, column=0, padx=5, pady=10, sticky="nsew")

    def delete(self) -> None:
        """Attempt to delete selected item and save back to excel.
        Also reschedule the bids.

        """
        selected_item = self.treeview.selection()[0]
        selected = self.treeview.index(self.treeview.selection())

        try:
            path = "bid.xlsx"
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active
            sheet.delete_rows(selected + 2, 1)
            workbook.save(path)
        except ValueError:
            pass

        self.treeview.delete(selected_item)
        self.schedule_on_start()

    def remove_old_bid(self) -> None:
        """Remove expired bids."""
        path = "bid.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        list_values = list(sheet.values)
        for row, item in enumerate(list_values[1:]):
            if item[-1] < dt.now():
                sheet.delete_rows(row + 2, 1)
                workbook.save(path)

    def start_gui(self) -> None:
        """Start GUI"""
        self.root.mainloop()

    def load_data(self) -> None:
        """Load valumes into the bid history column."""
        sheet = self.load_path("bid.xlsx")

        list_values = list(sheet.values)
        for col_name in list_values[0]:
            self.treeview.heading(col_name, text=col_name)

        for value_tuple in list_values[1:]:
            self.treeview.insert("", tk.END, values=value_tuple)

    def bid_time(self) -> datetime:
        """Get the time before the bid is due to end."""
        item = self.item_id.get()
        list_values = self.default_or_user_path()

        ebay_bidding = EbayBidding(
            chrome_user=list_values[0],
            chrome_user_path=list_values[1],
            chrome_exe=list_values[2],
            ebay_item_number=f"{item}")
        bid_end = ebay_bidding.get_minute_delayed_bid_time()
        ebay_bidding.quit()
        return bid_end

    def bid(self, item, amount, seconds)  -> None:
        """Place a bid on the item. Throwing an error
        to the error log if item has been outbid.

        """
        list_values = self.default_or_user_path()

        ebay_bidding = EbayBidding(
            chrome_user=list_values[0],
            chrome_user_path=list_values[1],
            chrome_exe=list_values[2],
            ebay_item_number=f"{item}")

        seconds = int(seconds)
        ebay_bidding.get_minute_delayed_bid_time()
        ebay_bidding.click_submit_bid()
        ebay_bidding.insert_bid_amount(amount=amount)

        seconds_before_bid = ebay_bidding.review_inserted_amount(
            seconds=seconds)

        if not seconds_before_bid:
            self.error.config(text="You have been outbid")
            ebay_bidding.quit()
        else:
            self.scheduler.add_job(
                ebay_bidding.confirm_bid,
                "date", run_date=seconds_before_bid)
            self.scheduler.add_job(
                self.remove_old_bid, "date",
                run_date=(seconds_before_bid
                          + datetime.timedelta(seconds=seconds)))

    def schedule_on_start(self)  -> None:
        """Goes through all saved bids and adds
        them back to the scheduled bid time.

        """
        sheet = self.load_path("bid.xlsx")
        list_values = list(sheet.values)
        self.scheduler.remove_all_jobs()

        for item in list_values[1:]:
            self.scheduler.add_job(
                self.bid,
                "date",
                run_date=item[-1],
                args=[item[0], item[1], item[2]])

        if len(list_values[1:]) > 0:
            self.scheduler.print_jobs()

    def insert_row(self) -> None:
        """Save the inserted item row and
        insert it into the bid history column

        """
        item = self.item_id.get()
        bid_amount = self.amount.get()
        seconds = self.seconds_combobox.get()
        bid_end = self.bid_time()

        # Insert row into Excel sheet
        path = "bid.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = self.load_path("bid.xlsx")
        row_values = [item, bid_amount, seconds, bid_end]
        sheet.append(row_values)
        workbook.save(path)

        self.save_chrome_path_and_exe()

        # Insert row into treeview
        self.treeview.insert("", tk.END, values=row_values)
        self.schedule_on_start()

    def start_ebay_bidding_gui(self):
        self.remove_old_bid()
        self.set_frame()
        self.bid_border()
        self.id_row()
        self.amount_row()
        self.seconds_row()
        self.set_chrome_user_path()
        self.set_chrome_exe()
        self.set_chrome_profile()
        self.submit_bid()
        self.bid_history()
        self.load_history()
        self.delete_row()
        self.save_path()
        self.error_log()
        self.error_text()
        self.schedule_on_start()
        self.start_gui()


if __name__ == "__main__":
    EbayGui().start_ebay_bidding_gui()
