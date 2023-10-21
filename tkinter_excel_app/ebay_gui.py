import tkinter as tk
from tkinter import ttk
import openpyxl
from bidder_class import EbayBidding
from apscheduler.schedulers.background import BackgroundScheduler

scheduler = BackgroundScheduler()
scheduler.start()

def load_data():
    path = "bid.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    list_values = list(sheet.values)
    for col_name in list_values[0]:
        treeview.heading(col_name, text=col_name)

    for value_tuple in list_values[1:]:
        treeview.insert("", tk.END, values=value_tuple)

def bid_time():
    item = item_id.get()

    ebay_bidding = EbayBidding(
        chrome_user="Default",
        chrome_user_path="C:\\Users\\a_asf\\AppData\\Local\\Google\\Chrome\\User Data\\",
        chrome_exe="C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe",
        ebay_item_number=f"{item}")
    bid_end = ebay_bidding.get_minute_delayed_bid_time()
    ebay_bidding.quit()
    return bid_end

def bid(item, amount, seconds):
    ebay_bidding = EbayBidding(
        chrome_user="Default",
        chrome_user_path="C:\\Users\\a_asf\\AppData\\Local\\Google\\Chrome\\User Data\\",
        chrome_exe="C:\\Program Files\\Google\\Chrome\\Application\\chrome.exe",
        ebay_item_number=f"{item}")
    
    seconds = int(seconds)
    ebay_bidding.get_minute_delayed_bid_time()
    seconds_before_bid = ebay_bidding.place_bid(seconds=seconds, amount=amount)
    print(seconds_before_bid)
    print(type(seconds_before_bid))
    scheduler.add_job(ebay_bidding.confirm_bid, "date", run_date=seconds_before_bid)


def schedule_on_start():
    path = "bid.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    list_values = list(sheet.values)
    scheduler.remove_all_jobs()
    
    for item in list_values[1:]:
        scheduler.add_job(bid, "date", run_date=item[-1], args=[item[0], item[1], item[2]])
    if len(list_values[1:]) > 0:
        scheduler.print_jobs()



def insert_row():
    item = item_id.get()
    bid_amount = amount.get()
    seconds = seconds_combobox.get()
    bid_end = bid_time()

    # Insert row into Excel sheet
    path = "bid.xlsx"
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    row_values = [item, bid_amount, seconds, bid_end]
    sheet.append(row_values)
    workbook.save(path)

    # Insert row into treeview
    treeview.insert("", tk.END, values=row_values)
    schedule_on_start()

root = tk.Tk()

style = ttk.Style(root)
root.tk.call("source", "forest-light.tcl")
root.tk.call("source", "forest-dark.tcl")
style.theme_use("forest-dark")

combo_list = [5, 4, 3, 2, 1]

frame = ttk.Frame(root)
frame.pack()

widgets_frame = ttk.LabelFrame(frame, text="Insert Bid")
widgets_frame.grid(row=0, column=0, sticky="nsew")

# Item ID row
item_id = ttk.Entry(widgets_frame)
item_id.insert(0, "Item ID")
item_id.bind("<FocusIn>", lambda e: item_id.delete("0", "end"))
item_id.grid(row=0, column=0, padx=5, pady=10, sticky="ew")

# Amount row
amount = ttk.Entry(widgets_frame)
amount.insert(0, "£")
amount.bind("<FocusIn>", lambda e: amount.delete("0", "end"))
amount.grid(row=1, column=0, padx=5, pady=10, sticky="ew")

# Seconds row
seconds_combobox = ttk.Combobox(widgets_frame, values=combo_list)
seconds_combobox.current(0)
seconds_combobox.grid(row=2, column=0, padx=5, pady=10, sticky="ew")

# Submit bid
button = ttk.Button(widgets_frame, text="Bid", command=insert_row)
button.grid(row=4, column=0, padx=5, pady=10, sticky="nsew")

# Bid history
treeFrame = ttk.Frame(frame)
treeFrame.grid(row=0, column=1)
treeScroll = ttk.Scrollbar(treeFrame)
treeScroll.pack(side="right", fill="y")

cols = ("Item ID", "Amount £", "Seconds", "Bid end")
treeview = ttk.Treeview(treeFrame, show="headings", yscrollcommand=treeScroll.set, columns=cols, height=13)
treeview.column("Item ID", width=100)
treeview.column("Amount £", width=50)
treeview.column("Seconds", width=100, anchor="center")
treeview.column("Bid end", width=200)
treeview.pack()
treeScroll.config(command=treeview.yview)
load_data()

# Delete row
delete_button = tk.Button(widgets_frame, text="Delete row", command=lambda:delete())
delete_button.grid(row=5, column=0, padx=5, pady=10, sticky="nsew")

def delete():
    selected_item=treeview.selection()[0]
    selected = treeview.index(treeview.selection())
    delete_item_id = treeview.item(treeview.selection(), "values")[0]

    try:
        path = "bid.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        sheet.delete_rows(selected + 2, 1)
        workbook.save(path)
    except:
        pass
    treeview.delete(selected_item)
    schedule_on_start()

root.mainloop()
